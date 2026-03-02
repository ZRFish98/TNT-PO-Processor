"""Payment Reconciler — Parse T&T cheque remittances and match against Odoo."""

from __future__ import annotations

import base64
import dataclasses
import io
import logging
from typing import TYPE_CHECKING, Dict, List, Optional

import pandas as pd
import streamlit as st

from backend.payment_extractor import PaymentStatement
from frontend.i18n import t

if TYPE_CHECKING:
    from backend.odoo_client import OdooClient

logger = logging.getLogger(__name__)

# Internal tab keys (language-independent)
_TAB_KEYS = ["upload", "reconcile", "create_payments"]

_STATUS_ICONS = {
    "matched": "\u2705",
    "amount_mismatch": "\u26a0\ufe0f",
    "not_found": "\u274c",
    "pending": "\u23f3",
}

_TYPE_LABELS_KEY = {
    "invoice": "payment.type.invoice",
    "credit_note": "payment.type.credit_note",
    "return": "payment.type.return",
}

_STATUS_LABELS_KEY = {
    "matched": "payment.status.matched",
    "amount_mismatch": "payment.status.amount_mismatch",
    "not_found": "payment.status.not_found",
    "pending": "payment.status.pending",
}

_PAYMENT_STATE_LABELS = {
    "not_paid": "Not Paid",
    "partial": "Partial",
    "in_payment": "In Payment",
    "paid": "Paid",
    "reversed": "Reversed",
}


# ── Tab 1: Upload & Extract ─────────────────────────────────────────────────


def _render_upload(settings: Dict) -> None:
    """Tab 1: Upload payment statement PDF and extract line items."""
    st.caption(t("payment.upload.caption"))

    uploaded = st.file_uploader(
        t("payment.upload.label"),
        type=["pdf"],
        key="payment_pdf_uploader",
    )

    if uploaded:
        # Only parse on new upload
        current_name = uploaded.name
        prev_name = st.session_state.get("_payment_uploaded_name")

        if current_name != prev_name:
            from backend.payment_extractor import PaymentExtractor

            pdf_bytes: bytes = uploaded.read()

            store_names = settings.get("tt_store_names", {})
            store_codes = settings.get("payment_store_codes", {})
            # settings.yaml keys are ints for store_names, strings for store_codes
            store_codes_str = {str(k): v for k, v in store_codes.items()}

            try:
                statement = PaymentExtractor.extract(
                    pdf_bytes, store_names, store_codes_str
                )
                st.session_state["payment_pdf_bytes"] = pdf_bytes
                st.session_state["payment_statement"] = statement
                st.session_state["payment_matched"] = False
                st.session_state["_payment_uploaded_name"] = current_name
            except Exception as e:
                logger.exception("Failed to extract payment statement")
                st.error(t("payment.upload.error", error=str(e)))
                return

    statement = st.session_state.get("payment_statement")
    if not statement:
        return

    # Summary metrics
    invoices = [ln for ln in statement.lines if ln.doc_type == "invoice"]
    credits = [ln for ln in statement.lines if ln.doc_type != "invoice"]

    st.success(
        t(
            "payment.upload.done",
            lines=len(statement.lines),
            pages=statement.page_count,
        )
    )

    c1, c2, c3 = st.columns(3)
    c1.metric(t("payment.upload.metric.cheque"), statement.cheque_number)
    c2.metric(t("payment.upload.metric.date"), statement.cheque_date)
    c3.metric(
        t("payment.upload.metric.total"), f"${statement.total_amount:,.2f}"
    )
    c4, c5, c6 = st.columns(3)
    c4.metric(t("payment.upload.metric.lines"), len(statement.lines))
    c5.metric(t("payment.upload.metric.invoices"), len(invoices))
    c6.metric(t("payment.upload.metric.credits"), len(credits))

    # PDF preview
    pdf_bytes = st.session_state.get("payment_pdf_bytes")
    if pdf_bytes:
        with st.expander(t("payment.upload.pdf_preview"), expanded=False):
            b64 = base64.b64encode(pdf_bytes).decode()
            pdf_html = (
                f'<iframe src="data:application/pdf;base64,{b64}" '
                f'width="100%" height="600" type="application/pdf"></iframe>'
            )
            st.markdown(pdf_html, unsafe_allow_html=True)

    # Preview table
    rows = []
    for ln in statement.lines:
        amount_display = -ln.net_amount if ln.is_credit else ln.net_amount
        rows.append(
            {
                t("payment.upload.col.store"): f"{ln.store_code}-{ln.store_id} {ln.store_name}",
                t("payment.upload.col.reference"): ln.reference_number,
                t("payment.upload.col.type"): t(_TYPE_LABELS_KEY.get(ln.doc_type, "payment.type.invoice")),
                t("payment.upload.col.date"): ln.invoice_date,
                t("payment.upload.col.amount"): amount_display,
            }
        )

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Navigation
    col_next, col_clear = st.columns([1, 1])

    with col_next:
        def _go_to_reconcile() -> None:
            st.session_state["payment_active_tab"] = "reconcile"
            st.session_state["_payment_tab_radio"] = t("payment.tab.reconcile")

        st.button(
            t("payment.upload.btn_next"),
            key="btn_payment_next_reconcile",
            on_click=_go_to_reconcile,
            type="primary",
        )

    with col_clear:
        if st.button(t("payment.upload.btn_clear"), key="btn_payment_clear"):
            for k in (
                "payment_pdf_bytes",
                "payment_statement",
                "payment_matched",
                "payment_matched_lines",
                "_payment_uploaded_name",
                "payment_bank_journals",
                "payment_batches_created",
                "payment_batch_results",
            ):
                st.session_state.pop(k, None)
            st.rerun()


# ── Tab 2: Match & Reconcile ────────────────────────────────────────────────


def _match_with_odoo(
    statement: PaymentStatement,
    client: OdooClient,
) -> List[Dict[str, object]]:
    """Match all payment lines against Odoo account.move records.

    Returns a list of dicts (one per line) with original fields plus Odoo
    match info (``odoo_move_id``, ``odoo_amount``, ``match_status``, etc.).
    """

    # 1. Collect references by type
    invoice_refs = list(
        {ln.odoo_search_ref for ln in statement.lines if ln.doc_type == "invoice"}
    )
    credit_refs = list(
        {ln.odoo_search_ref for ln in statement.lines if ln.doc_type == "credit_note"}
    )
    return_refs = list(
        {ln.odoo_search_ref for ln in statement.lines if ln.doc_type == "return"}
    )

    # 2. Batch search Odoo
    inv_results = client.search_invoices_by_name(invoice_refs) if invoice_refs else []
    cn_results = (
        client.search_account_moves_by_pattern(credit_refs, move_type="out_refund")
        if credit_refs
        else []
    )
    ret_results = (
        client.search_account_moves_by_pattern(return_refs) if return_refs else []
    )

    # 3. Build lookup maps: search_ref -> best matching record
    inv_lookup: Dict[str, dict] = {}
    for r in inv_results:
        name = r.get("name", "") or ""
        for ref in invoice_refs:
            if ref in name and ref not in inv_lookup:
                inv_lookup[ref] = r

    cn_lookup: Dict[str, dict] = {}
    for r in cn_results:
        name = r.get("name", "") or ""
        ref_field = r.get("ref", "") or ""
        for cn_ref in credit_refs:
            if (cn_ref in name or cn_ref in ref_field) and cn_ref not in cn_lookup:
                cn_lookup[cn_ref] = r

    ret_lookup: Dict[str, dict] = {}
    for r in ret_results:
        name = r.get("name", "") or ""
        ref_field = r.get("ref", "") or ""
        for ret_ref in return_refs:
            if (ret_ref in name or ret_ref in ref_field) and ret_ref not in ret_lookup:
                ret_lookup[ret_ref] = r

    # 4. Update each line
    updated_lines: List[dict] = []
    for ln in statement.lines:
        match: Optional[dict] = None
        if ln.doc_type == "invoice":
            match = inv_lookup.get(ln.odoo_search_ref)
        elif ln.doc_type == "credit_note":
            match = cn_lookup.get(ln.odoo_search_ref)
        elif ln.doc_type == "return":
            match = ret_lookup.get(ln.odoo_search_ref)

        ln_dict = dataclasses.asdict(ln)

        if match:
            odoo_amount = match.get("amount_total", 0)
            amount_match = abs(ln.net_amount - abs(odoo_amount)) < 0.02

            partner = match.get("partner_id")
            partner_name = (
                partner[1] if isinstance(partner, (list, tuple)) else str(partner or "")
            )

            partner_id = (
                partner[0] if isinstance(partner, (list, tuple)) else partner
            )

            ln_dict["odoo_move_id"] = match["id"]
            ln_dict["odoo_move_name"] = match.get("name", "")
            ln_dict["odoo_amount"] = odoo_amount
            ln_dict["odoo_payment_state"] = match.get("payment_state", "")
            ln_dict["odoo_partner_name"] = partner_name
            ln_dict["odoo_partner_id"] = partner_id
            ln_dict["match_status"] = "matched" if amount_match else "amount_mismatch"
        else:
            ln_dict["odoo_move_id"] = None
            ln_dict["odoo_move_name"] = None
            ln_dict["odoo_amount"] = None
            ln_dict["odoo_payment_state"] = None
            ln_dict["odoo_partner_name"] = None
            ln_dict["odoo_partner_id"] = None
            ln_dict["match_status"] = "not_found"

        updated_lines.append(ln_dict)

    return updated_lines


def _render_reconcile(settings: Dict) -> None:
    """Tab 2: Match extracted lines with Odoo and display results."""
    statement = st.session_state.get("payment_statement")
    if not statement:
        st.info(t("payment.reconcile.no_data"))
        return

    client = st.session_state.get("odoo_client")
    if not client or not st.session_state.get("odoo_connected"):
        st.error(t("payment.reconcile.no_odoo"))
        return

    # Match button
    if st.button(
        t("payment.reconcile.btn_match"), type="primary", key="btn_payment_match"
    ):
        with st.spinner(
            t("payment.reconcile.matching", count=len(statement.lines))
        ):
            matched_lines = _match_with_odoo(statement, client)
            st.session_state["payment_matched_lines"] = matched_lines
            st.session_state["payment_matched"] = True
        st.rerun()

    if not st.session_state.get("payment_matched"):
        return

    matched_lines: List[dict] = st.session_state.get("payment_matched_lines", [])
    if not matched_lines:
        return

    # Compute stats
    total = len(matched_lines)
    matched_count = sum(1 for ln in matched_lines if ln["match_status"] == "matched")
    mismatch_count = sum(
        1 for ln in matched_lines if ln["match_status"] == "amount_mismatch"
    )
    not_found_count = sum(
        1 for ln in matched_lines if ln["match_status"] == "not_found"
    )

    pdf_total = sum(
        (-ln["net_amount"] if ln["is_credit"] else ln["net_amount"])
        for ln in matched_lines
    )
    odoo_total = sum(
        (
            -abs(ln["odoo_amount"])
            if ln["is_credit"] and ln["odoo_amount"]
            else (ln["odoo_amount"] or 0)
        )
        for ln in matched_lines
    )
    variance = pdf_total - odoo_total

    # Summary message
    st.info(
        t(
            "payment.reconcile.done",
            matched=matched_count,
            total=total,
            mismatches=mismatch_count,
            not_found=not_found_count,
        )
    )

    # Summary metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(t("payment.reconcile.metric.total_lines"), total)
    c2.metric(t("payment.reconcile.metric.matched"), matched_count)
    c3.metric(t("payment.reconcile.metric.mismatched"), mismatch_count)
    c4.metric(t("payment.reconcile.metric.not_found"), not_found_count)

    c5, c6, c7 = st.columns(3)
    c5.metric(t("payment.reconcile.metric.pdf_total"), f"${pdf_total:,.2f}")
    c6.metric(t("payment.reconcile.metric.odoo_total"), f"${odoo_total:,.2f}")
    c7.metric(
        t("payment.reconcile.metric.variance"),
        f"${variance:,.2f}",
        delta=f"${variance:,.2f}" if abs(variance) > 0.01 else None,
        delta_color="inverse",
    )

    st.divider()

    # Refresh button — re-match after user edits in Odoo
    if st.button(
        t("payment.reconcile.btn_refresh"),
        key="btn_payment_refresh",
        icon=":material/refresh:",
    ):
        with st.spinner(
            t("payment.reconcile.matching", count=len(statement.lines))
        ):
            matched_lines = _match_with_odoo(statement, client)
            st.session_state["payment_matched_lines"] = matched_lines
            st.session_state["payment_matched"] = True
            # Reset batch creation guard since data changed
            st.session_state.pop("payment_batches_created", None)
            st.session_state.pop("payment_batch_results", None)
        st.rerun()

    # Filters
    fc1, fc2 = st.columns(2)

    stores = sorted({ln["store_code"] + "-" + ln["store_id"] + " " + ln["store_name"] for ln in matched_lines})
    with fc1:
        store_filter = st.selectbox(
            t("payment.reconcile.filter.store"),
            [t("payment.reconcile.filter.store_all")] + stores,
            key="payment_store_filter",
        )

    status_options = {
        t("payment.reconcile.filter.status_all"): None,
        t("payment.reconcile.filter.status_matched"): "matched",
        t("payment.reconcile.filter.status_mismatch"): "amount_mismatch",
        t("payment.reconcile.filter.status_not_found"): "not_found",
    }
    with fc2:
        status_label = st.selectbox(
            t("payment.reconcile.filter.status"),
            list(status_options.keys()),
            key="payment_status_filter",
        )
    status_filter = status_options.get(status_label)

    # Apply filters
    filtered = matched_lines
    if store_filter != t("payment.reconcile.filter.store_all"):
        store_code_filter = store_filter.split("-")[0]
        filtered = [ln for ln in filtered if ln["store_code"] == store_code_filter]
    if status_filter:
        filtered = [ln for ln in filtered if ln["match_status"] == status_filter]

    # Build display DataFrame
    rows = []
    for ln in filtered:
        status_icon = _STATUS_ICONS.get(ln["match_status"], "")
        status_text = t(_STATUS_LABELS_KEY.get(ln["match_status"], "payment.status.pending"))
        type_text = t(_TYPE_LABELS_KEY.get(ln["doc_type"], "payment.type.invoice"))

        pdf_amount = -ln["net_amount"] if ln["is_credit"] else ln["net_amount"]
        odoo_amount = ln["odoo_amount"]
        if odoo_amount is not None and ln["is_credit"]:
            odoo_display = -abs(odoo_amount)
        else:
            odoo_display = odoo_amount

        var = (pdf_amount - odoo_display) if odoo_display is not None else None

        payment_state = ln.get("odoo_payment_state", "") or ""
        payment_display = _PAYMENT_STATE_LABELS.get(payment_state, payment_state)

        rows.append(
            {
                t("payment.reconcile.col.status"): f"{status_icon} {status_text}",
                t("payment.reconcile.col.store"): f"{ln['store_code']}-{ln['store_id']} {ln['store_name']}",
                t("payment.reconcile.col.reference"): ln["reference_number"],
                t("payment.reconcile.col.doc_ref"): ln["odoo_search_ref"],
                t("payment.reconcile.col.type"): type_text,
                t("payment.reconcile.col.invoice_date"): ln["invoice_date"],
                t("payment.reconcile.col.pdf_amount"): pdf_amount,
                t("payment.reconcile.col.odoo_amount"): odoo_display if odoo_display is not None else "",
                t("payment.reconcile.col.variance"): var if var is not None else "",
                t("payment.reconcile.col.odoo_name"): ln.get("odoo_move_name") or "",
                t("payment.reconcile.col.payment_state"): payment_display,
            }
        )

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True, height=600)

    # Unmatched items expander
    not_found = [ln for ln in matched_lines if ln["match_status"] == "not_found"]
    if not_found:
        with st.expander(
            t("payment.reconcile.unmatched_title", count=len(not_found)),
            expanded=True,
        ):
            nf_rows = []
            for ln in not_found:
                nf_rows.append(
                    {
                        t("payment.reconcile.col.store"): f"{ln['store_code']}-{ln['store_id']}",
                        t("payment.reconcile.col.reference"): ln["reference_number"],
                        t("payment.reconcile.col.doc_ref"): ln["odoo_search_ref"],
                        t("payment.reconcile.col.type"): t(_TYPE_LABELS_KEY.get(ln["doc_type"], "")),
                        t("payment.reconcile.col.invoice_date"): ln["invoice_date"],
                        t("payment.reconcile.col.pdf_amount"): -ln["net_amount"] if ln["is_credit"] else ln["net_amount"],
                    }
                )
            st.dataframe(pd.DataFrame(nf_rows), use_container_width=True, hide_index=True)

    # Mismatches expander
    mismatches = [ln for ln in matched_lines if ln["match_status"] == "amount_mismatch"]
    if mismatches:
        with st.expander(
            t("payment.reconcile.mismatch_title", count=len(mismatches))
        ):
            mm_rows = []
            for ln in mismatches:
                pdf_amt = -ln["net_amount"] if ln["is_credit"] else ln["net_amount"]
                odoo_amt = ln["odoo_amount"] or 0
                if ln["is_credit"]:
                    odoo_amt = -abs(odoo_amt)
                mm_rows.append(
                    {
                        t("payment.reconcile.col.reference"): ln["reference_number"],
                        t("payment.reconcile.col.odoo_name"): ln.get("odoo_move_name", ""),
                        t("payment.reconcile.col.pdf_amount"): pdf_amt,
                        t("payment.reconcile.col.odoo_amount"): odoo_amt,
                        t("payment.reconcile.col.variance"): pdf_amt - odoo_amt,
                    }
                )
            st.dataframe(pd.DataFrame(mm_rows), use_container_width=True, hide_index=True)

    # Excel download
    st.divider()
    excel_bytes = _generate_excel(matched_lines, statement)
    st.download_button(
        label=t("payment.reconcile.btn_download"),
        data=excel_bytes,
        file_name=f"payment_reconciliation_{statement.cheque_number}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="btn_payment_download_file",
    )


# ── Tab 3: Create Payments ─────────────────────────────────────────────────


def _parse_cheque_date(raw_date: str) -> str:
    """Convert 'Feb 01, 26' to '2026-02-01' for Odoo."""
    from dateutil import parser as _dp

    try:
        dt = _dp.parse(raw_date, dayfirst=False)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return raw_date


def _render_create_payments(settings: Dict) -> None:
    """Tab 3: Create batch payments in Odoo from matched lines."""
    statement: PaymentStatement | None = st.session_state.get("payment_statement")
    if not statement or not st.session_state.get("payment_matched"):
        st.info(t("payment.create.no_data"))
        return

    client = st.session_state.get("odoo_client")
    if not client or not st.session_state.get("odoo_connected"):
        st.error(t("payment.create.no_odoo"))
        return

    matched_lines: List[dict] = st.session_state.get("payment_matched_lines", [])

    # Filter to payable lines (matched or amount_mismatch with an Odoo move ID)
    payable = [
        ln for ln in matched_lines
        if ln.get("match_status") in ("matched", "amount_mismatch")
        and ln.get("odoo_move_id")
    ]
    if not payable:
        st.warning(t("payment.create.no_payable"))
        return

    inbound_lines = [ln for ln in payable if not ln["is_credit"]]
    outbound_lines = [ln for ln in payable if ln["is_credit"]]

    inbound_total = sum(ln["net_amount"] for ln in inbound_lines)
    outbound_total = sum(ln["net_amount"] for ln in outbound_lines)

    # Summary metrics
    c1, c2 = st.columns(2)
    c1.metric(
        t("payment.create.metric.inbound"),
        f"{len(inbound_lines)} — ${inbound_total:,.2f}",
    )
    c2.metric(
        t("payment.create.metric.outbound"),
        f"{len(outbound_lines)} — ${outbound_total:,.2f}",
    )

    # Cheque info
    st.info(
        t(
            "payment.create.cheque_info",
            number=statement.cheque_number,
            date=statement.cheque_date,
        )
    )

    # Journal selection
    if st.session_state.get("payment_bank_journals") is None:
        with st.spinner(t("payment.create.loading_journals")):
            journals = client.get_bank_journals()
            st.session_state["payment_bank_journals"] = journals

    journals: list = st.session_state.get("payment_bank_journals", [])
    if not journals:
        st.error(t("payment.create.no_journals"))
        return

    journal_options = {j["name"]: j["id"] for j in journals}
    selected_name = st.selectbox(
        t("payment.create.journal_label"),
        list(journal_options.keys()),
        key="payment_journal_select",
    )
    journal_id = journal_options[selected_name]

    # Preview expanders
    if inbound_lines:
        with st.expander(
            t("payment.create.inbound_title", count=len(inbound_lines)),
            expanded=False,
        ):
            in_rows = [
                {
                    t("payment.upload.col.reference"): ln["reference_number"],
                    t("payment.reconcile.col.odoo_name"): ln.get("odoo_move_name", ""),
                    t("payment.upload.col.amount"): ln["net_amount"],
                }
                for ln in inbound_lines
            ]
            st.dataframe(pd.DataFrame(in_rows), use_container_width=True, hide_index=True)

    if outbound_lines:
        with st.expander(
            t("payment.create.outbound_title", count=len(outbound_lines)),
            expanded=False,
        ):
            out_rows = [
                {
                    t("payment.upload.col.reference"): ln["reference_number"],
                    t("payment.reconcile.col.odoo_name"): ln.get("odoo_move_name", ""),
                    t("payment.upload.col.amount"): ln["net_amount"],
                }
                for ln in outbound_lines
            ]
            st.dataframe(pd.DataFrame(out_rows), use_container_width=True, hide_index=True)

    st.divider()

    # Idempotency guard
    if st.session_state.get("payment_batches_created"):
        st.warning(t("payment.create.already_created"))
        results = st.session_state.get("payment_batch_results")
        if results:
            st.success(
                t(
                    "payment.create.done",
                    inbound=results.get("inbound_count", 0),
                    outbound=results.get("outbound_count", 0),
                    batches=results.get("batch_count", 0),
                )
            )
        return

    # Create button
    if st.button(
        t("payment.create.btn_create", count=len(payable)),
        type="primary",
        key="btn_create_batch_payments",
    ):
        _execute_batch_creation(
            client, statement, inbound_lines, outbound_lines, journal_id
        )


def _execute_batch_creation(
    client: OdooClient,
    statement: PaymentStatement,
    inbound_lines: List[dict],
    outbound_lines: List[dict],
    journal_id: int,
) -> None:
    """Run the full batch payment creation workflow."""
    cheque_date = _parse_cheque_date(statement.cheque_date)
    total_count = len(inbound_lines) + len(outbound_lines)

    with st.spinner(t("payment.create.spinner", count=total_count)):
        # 1. Find payment method lines
        in_method = client.get_payment_method_line(journal_id, "inbound")
        out_method = client.get_payment_method_line(journal_id, "outbound")
        if not in_method or not out_method:
            st.warning(t("payment.create.method_warn"))

        # 2. Create inbound payments (invoices) — use per-line partner_id
        inbound_ids: List[int] = []
        for ln in inbound_lines:
            line_partner = ln.get("odoo_partner_id")
            if not line_partner:
                logger.error(
                    "Skipping %s — no partner_id resolved", ln["reference_number"]
                )
                st.error(
                    t(
                        "payment.create.payment_fail",
                        ref=ln["reference_number"],
                        error="no partner_id",
                    )
                )
                continue
            vals = {
                "payment_type": "inbound",
                "partner_type": "customer",
                "partner_id": line_partner,
                "journal_id": journal_id,
                "amount": ln["net_amount"],
                "date": cheque_date,
                "ref": f"CHQ {statement.cheque_number} - {ln.get('odoo_move_name', ln['reference_number'])}",
            }
            if in_method:
                vals["payment_method_line_id"] = in_method
            pid = client.create_payment(vals)
            if pid:
                inbound_ids.append(pid)
            else:
                st.error(
                    t(
                        "payment.create.payment_fail",
                        ref=ln["reference_number"],
                        error="create returned None",
                    )
                )

        # 3. Create outbound payments (credits/returns) — use per-line partner_id
        outbound_ids: List[int] = []
        for ln in outbound_lines:
            line_partner = ln.get("odoo_partner_id")
            if not line_partner:
                logger.error(
                    "Skipping %s — no partner_id resolved", ln["reference_number"]
                )
                st.error(
                    t(
                        "payment.create.payment_fail",
                        ref=ln["reference_number"],
                        error="no partner_id",
                    )
                )
                continue
            vals = {
                "payment_type": "outbound",
                "partner_type": "customer",
                "partner_id": line_partner,
                "journal_id": journal_id,
                "amount": ln["net_amount"],
                "date": cheque_date,
                "ref": f"CHQ {statement.cheque_number} - {ln.get('odoo_move_name', ln['reference_number'])}",
            }
            if out_method:
                vals["payment_method_line_id"] = out_method
            pid = client.create_payment(vals)
            if pid:
                outbound_ids.append(pid)
            else:
                st.error(
                    t(
                        "payment.create.payment_fail",
                        ref=ln["reference_number"],
                        error="create returned None",
                    )
                )

    # 5. Create batch payment records
    batch_count = 0
    batch_in_id = None
    batch_out_id = None

    if inbound_ids:
        with st.spinner(
            t(
                "payment.create.spinner_batch",
                name=f"CHQ {statement.cheque_number} — Invoices",
            )
        ):
            batch_in_id = client.create_batch_payment(
                name=f"CHQ {statement.cheque_number} — Invoices",
                journal_id=journal_id,
                payment_ids=inbound_ids,
            )
            if batch_in_id:
                batch_count += 1

    if outbound_ids:
        with st.spinner(
            t(
                "payment.create.spinner_batch",
                name=f"CHQ {statement.cheque_number} — Credits & Returns",
            )
        ):
            batch_out_id = client.create_batch_payment(
                name=f"CHQ {statement.cheque_number} — Credits & Returns",
                journal_id=journal_id,
                payment_ids=outbound_ids,
            )
            if batch_out_id:
                batch_count += 1

    # 6. Attach PDF to batches
    pdf_bytes = st.session_state.get("payment_pdf_bytes")
    if pdf_bytes:
        filename = f"Statement_CHQ_{statement.cheque_number}.pdf"
        with st.spinner(t("payment.create.spinner_attach")):
            for batch_id in (batch_in_id, batch_out_id):
                if not batch_id:
                    continue
                att_id = client.attach_pdf_to_record(
                    "account.batch.payment", batch_id, filename, pdf_bytes
                )
                if att_id:
                    client.post_chatter_message(
                        "account.batch.payment",
                        batch_id,
                        f"Payment statement for cheque #{statement.cheque_number}",
                        attachment_ids=[att_id],
                    )

    # 7. Store results + guard
    results = {
        "inbound_count": len(inbound_ids),
        "outbound_count": len(outbound_ids),
        "batch_count": batch_count,
        "batch_in_id": batch_in_id,
        "batch_out_id": batch_out_id,
    }
    st.session_state["payment_batches_created"] = True
    st.session_state["payment_batch_results"] = results

    st.success(
        t(
            "payment.create.done",
            inbound=len(inbound_ids),
            outbound=len(outbound_ids),
            batches=batch_count,
        )
    )


def _generate_excel(
    matched_lines: List[dict],
    statement: PaymentStatement,
) -> bytes:
    """Generate reconciliation Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Payment Reconciliation Report"])
    ws_summary.append(["Cheque #", statement.cheque_number])
    ws_summary.append(["Date", statement.cheque_date])
    ws_summary.append(["Total Amount", statement.total_amount])
    ws_summary.append([])

    # Summary by store
    ws_summary.append(["Store", "Invoices", "Credits", "Total", "Matched", "Not Found"])
    stores: Dict[str, Dict[str, float]] = {}
    for ln in matched_lines:
        key = f"{ln['store_code']}-{ln['store_id']} {ln['store_name']}"
        if key not in stores:
            stores[key] = {"invoices": 0, "credits": 0, "total": 0.0, "matched": 0, "not_found": 0}
        s = stores[key]
        if ln["doc_type"] == "invoice":
            s["invoices"] += 1
        else:
            s["credits"] += 1
        amt = -ln["net_amount"] if ln["is_credit"] else ln["net_amount"]
        s["total"] += amt
        if ln["match_status"] in ("matched", "amount_mismatch"):
            s["matched"] += 1
        elif ln["match_status"] == "not_found":
            s["not_found"] += 1

    for store_name, s in sorted(stores.items()):
        ws_summary.append([
            store_name,
            s["invoices"],
            s["credits"],
            round(s["total"], 2),
            s["matched"],
            s["not_found"],
        ])

    # Sheet 2: Details
    ws_detail = wb.create_sheet("Details")
    headers = [
        "Status", "Store", "Reference", "Doc Reference", "Type",
        "Invoice Date", "PDF Amount", "Odoo Amount", "Variance",
        "Odoo Doc Name", "Payment State",
    ]
    ws_detail.append(headers)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    for ln in matched_lines:
        pdf_amt = -ln["net_amount"] if ln["is_credit"] else ln["net_amount"]
        odoo_amt = ln["odoo_amount"]
        if odoo_amt is not None and ln["is_credit"]:
            odoo_amt = -abs(odoo_amt)
        var = (pdf_amt - odoo_amt) if odoo_amt is not None else None

        status_map = {
            "matched": "Matched",
            "amount_mismatch": "Amount Mismatch",
            "not_found": "NOT FOUND",
            "pending": "Pending",
        }

        row_data = [
            status_map.get(ln["match_status"], ln["match_status"]),
            f"{ln['store_code']}-{ln['store_id']} {ln['store_name']}",
            ln["reference_number"],
            ln["odoo_search_ref"],
            ln["doc_type"].replace("_", " ").title(),
            ln["invoice_date"],
            pdf_amt,
            odoo_amt if odoo_amt is not None else "",
            round(var, 2) if var is not None else "",
            ln.get("odoo_move_name") or "",
            _PAYMENT_STATE_LABELS.get(ln.get("odoo_payment_state", ""), ""),
        ]
        ws_detail.append(row_data)

        # Highlight row
        row_num = ws_detail.max_row
        if ln["match_status"] == "not_found":
            for cell in ws_detail[row_num]:
                cell.fill = red_fill
        elif ln["match_status"] == "amount_mismatch":
            for cell in ws_detail[row_num]:
                cell.fill = yellow_fill

    # Bold headers
    bold = Font(bold=True)
    for cell in ws_detail[1]:
        cell.font = bold
    for cell in ws_summary[1]:
        cell.font = bold

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Main render ──────────────────────────────────────────────────────────────


def render(settings: Dict) -> None:
    """Main render function for Payment Reconciler."""
    st.title(t("payment.title"))

    if "payment_active_tab" not in st.session_state:
        st.session_state["payment_active_tab"] = _TAB_KEYS[0]

    _tab_labels = [
        t("payment.tab.upload"),
        t("payment.tab.reconcile"),
        t("payment.tab.create_payments"),
    ]
    _key_to_label = dict(zip(_TAB_KEYS, _tab_labels))
    _label_to_key = dict(zip(_tab_labels, _TAB_KEYS))

    if "_payment_tab_radio" not in st.session_state:
        current_key = st.session_state["payment_active_tab"]
        st.session_state["_payment_tab_radio"] = _key_to_label.get(
            current_key, _tab_labels[0]
        )

    active_label = st.radio(
        t("payment.workflow_step"),
        _tab_labels,
        horizontal=True,
        key="_payment_tab_radio",
        label_visibility="collapsed",
    )
    st.session_state["payment_active_tab"] = _label_to_key.get(
        active_label, _TAB_KEYS[0]
    )

    st.divider()

    active_key = st.session_state["payment_active_tab"]
    if active_key == "upload":
        _render_upload(settings)
    elif active_key == "reconcile":
        _render_reconcile(settings)
    elif active_key == "create_payments":
        _render_create_payments(settings)
